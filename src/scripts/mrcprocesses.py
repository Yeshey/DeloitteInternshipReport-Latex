        import json
        import pyautogui
        import pyperclip
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from bs4 import BeautifulSoup
        import re
        import time
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import pandas as pd
        from selenium.common.exceptions import WebDriverException
        from selenium.common.exceptions import WebDriverException, StaleElementReferenceException
        from selenium.webdriver.common.alert import Alert
        import openpyxl
        from openpyxl import load_workbook
        import subprocess
        import keyboard
        import toml
        
        # pip install -r requirements.txt
        
        running = True
        
        def stop_script():
            global running
            print("Script stopping on next iteration")
            running = False
            
        def install_dependencies():
            with open('requirements.txt', 'r') as file:
                dependencies = file.read().splitlines()
        
            for dependency in dependencies:
                subprocess.run(['pip', 'install', dependency])
        
        def find_element(driver, xpath, timeout=10, max_attempts=20):
            button = None
            try:
                button = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            except (WebDriverException, StaleElementReferenceException) as e:
                print(f"Catching WebDriverException: {e}")
                button = None
        
            attempts = 0
            while attempts < max_attempts and button is None:
                try:
                    button = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                except (WebDriverException, StaleElementReferenceException):
                    attempts += 1
                    pass
            else:
                if attempts == max_attempts and button is None:
                    raise Exception(f"Failed after {max_attempts} attempts")
            return button
        
        def initialize_driver():
            global driver
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
        
            driver = webdriver.Chrome(options)  # options
        
            with open('credentials.toml', 'r') as file:
                credentials = toml.load(file)['credentials']
        
            process_id = "2101221"
            first_url = f"https://www.ril.com/servicecenter/Process_Instance_Detail.aspx?ProcessId={process_id}"
            driver.get(first_url)
        
            second_url = "https://cloud.mongodb.com/v2/metrics/replicaSet/explorer/application/users/find"
            driver.execute_script(f"window.open('{second_url}', '_blank');")
        
            # login into service center
            driver.switch_to.window(driver.window_handles[0])
            driver.implicitly_wait(3)
        
            # Find the username and password fields and fill them
            username_field = driver.find_element(By.ID, "wt89_wtContentRight_wtInput1")
            password_field = driver.find_element(By.ID, "wt89_wtContentRight_wtInputPass1")
        
            username_field.send_keys(credentials['username'])
            password_field.send_keys(credentials['password'])
        
            # Find and click the login button
            login_button = driver.find_element(By.ID, "wt89_wtContentRight_wt59_wtColumnsItems_wt33_wtContent_wtButton1")
            login_button.click()
        
            driver.switch_to.window(driver.window_handles[1])
            username_field = driver.find_element(By.ID, "username")
        
            username_field.send_keys(credentials['email'])
        
            el = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//button[@type='submit' and contains(., 'Next')]")))
            WebDriverWait(driver, 10).until(lambda d: 'false' in el.get_attribute('aria-disabled'))
            el.click()
            
            try:
                WebDriverWait(driver, 200).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[1]/div/form/div/div[2]/div/div/div/div/div/div/div[2]/div[1]")))
            except (WebDriverException, StaleElementReferenceException) as e:
                print(f"Catching WebDriverException: {e}")
                mongoloaded = None
                while mongoloaded is None:
                    try:
                        mongoloaded = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[1]/div/form/div/div[2]/div/div/div/div/div/div/div[2]/div[1]")))
                    except (WebDriverException, StaleElementReferenceException):
                        pass
            
        def get_process_info(process_id):
            # URL for the service center processes
            url = f"https://www.ril.com/servicecenter/Process_Instance_Detail.aspx?ProcessId={process_id}"
        
            driver.switch_to.window(driver.window_handles[0])
            driver.get(url)
        
            driver.implicitly_wait(5)
        
            target_div_id = "wt70_wtContentTop_wtListRecords1"
            target_div = driver.find_element(By.ID, target_div_id)
        
            html_content = target_div.get_attribute('innerHTML')
        
            soup = BeautifulSoup(html_content, 'html.parser')
            
            result = {}
        
            # Find all div elements with a specific class
            top_level_divs = soup.find_all('div', class_='columns gutter-base align-items-center margin-bottom-xs')
        
            # Loop through each div element
            for div in top_level_divs:
                # Find the span element inside the first div
                field_name_element = div.find('span', class_='text-neutral-7')
        
                # Find the content inside the second div
                second_div = div.find('div', class_='columns-item').find_next('div', class_='columns-item')
        
                # Check if both elements are present
                if field_name_element and second_div:
                    field_name = field_name_element.get_text(strip=True)
                    field_value = second_div.get_text(strip=True)
        
                    # Add the field and its corresponding value to the result dictionary
                    result[field_name] = field_value
                else:
                    print(f"div_elements:\n{top_level_divs}")
                    print(f"div:\n{div}")
                    print(f"html_content:\n{html_content}")
                    print(f"result:\n{result}")
                    raise Exception("Error getting some of the process info")
                
            return result
        
        def is_user_active(uwxid):
            driver.switch_to.window(driver.window_handles[1])
            time.sleep(1) # if it's too fast, it will desselect after clicking
            element = find_element(driver, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[1]/div/form/div/div[2]/div/div/div/div/div/div/div[2]/div[1]") 
            element.click()
            driver.implicitly_wait(1)
            # search
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('delete')
            text_query = '{_id: "' + uwxid + '"}'
            pyperclip.copy(text_query)
            pyautogui.hotkey('ctrl', 'v')
            driver.implicitly_wait(1)
            pyautogui.press('enter')
            
            # Wait for loading documents to appear and disappear
            try:
                WebDriverWait(driver, 5
                    ).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'data-explorer-subheader-info-is-loading') and contains(text(), 'Loading Documents')]")))
                WebDriverWait(driver, 10
                    ).until_not(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'data-explorer-subheader-info-is-loading') and contains(text(), 'Loading Documents')]")))
            except (WebDriverException, StaleElementReferenceException) as e:
                print(f"Catching WebDriverException: {e}")
                element = find_element(driver, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[1]/div/form/div/div[2]/div/div/div/div/div/div/div[2]/div[1]") 
                element.click()
                driver.implicitly_wait(1)
                # search
                pyautogui.hotkey('ctrl', 'a')
                pyautogui.press('delete')
                text_query = '{_id: "' + uwxid + '"}'
                pyperclip.copy(text_query)
                pyautogui.hotkey('ctrl', 'v')
                driver.implicitly_wait(1)
                pyautogui.press('enter')
                time.sleep(5) # alternativeley just wait 5 seconds
        
            # move mouse over query result for copy button to become visible
            query_result = find_element(driver, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[3]/div/div[2]") 
            ActionChains(driver).move_to_element(query_result).perform()
            driver.implicitly_wait(10)
            
            # copy the query result
            button = find_element(driver, "/html/body/div[1]/div/div/div[3]/div[1]/div[3]/main/div/div/div[2]/div[2]/div[2]/div[5]/div/div[3]/div/div[2]/div/div/div[2]/div/button") 
            driver.implicitly_wait(10)
            button.click()
            clipboard_content = pyperclip.paste()
            data = json.loads(clipboard_content) # Parse the JSON string
            active_value = data.get('active', None) # Extract the value of the 'active' key
        
            return active_value
        
        def fill_in_excel_openpyxl(sheet, row_index, process_data):
            # Extract relevant information from process_data
            contract_id = process_data.get('ContractId', None)
            placement_id = process_data.get('PlacementId', None)
            section_id = process_data.get('SectionId', None)
            submission_id = process_data.get('SubmissionId', None)
            uwx_id = process_data.get('UWXId', None)
        
            # Update the specified columns in the row using openpyxl
            sheet.cell(row=row_index + 2, column=2, value=contract_id)  # Second column for ContractId
            sheet.cell(row=row_index + 2, column=3, value=placement_id)  # Third column for PlacementId
            sheet.cell(row=row_index + 2, column=4, value=section_id)  # Fourth column for SectionId
            sheet.cell(row=row_index + 2, column=5, value=submission_id)  # Fifth column for SubmissionId
            sheet.cell(row=row_index + 2, column=7, value=uwx_id)  # Seventh column for UWXId
        
        def fill_excel_and_skip_process_openpyxl(sheet, row_index):
            skipped = True
            sheet.cell(row=row_index + 2, column=6, value="User not active")  # Seventh column for "User not active"
            
            # Skip the process automatically
            driver.switch_to.window(driver.window_handles[0])
            try:
                skip_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[3]/table/tbody/tr[1]/td[11]/input[2]")))
                driver.implicitly_wait(1)
                skip_button.click()
                WebDriverWait(driver, 10).until(EC.alert_is_present())
                time.sleep(0.5)
                driver.switch_to.alert.accept()
                time.sleep(3)
            except:
                print("Couldn't skip, saving without 'Y' in Skipped")
                skipped = False
                
            if skipped:
                sheet.cell(row=row_index + 2, column=8, value="Y")  # Ninth column for "Y"
            return skipped        
        
        if __name__ == "__main__":
            install_dependencies()
            
            # Load the Excel file with openpyxl
            excel_file = 'SDC_processes.xlsx'
            wb = load_workbook(excel_file)
            sheet = wb['GenerateMRCEMail']
            
            keyboard.add_hotkey('ctrl+alt+s', stop_script) # stop with CTRL + ALT + DELETE
            
            initialize_driver()
            try:
                # analyse ProcessIds in the first column from top to bottom if the second column is empty
                for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True)):
                    process_id = row[0]
                    if process_id is None: # we reached the end
                        print("No more processIds to analyse")
                        break
                    if running == False:
                        print("script stopping...")
                        break
                    second_column_value = row[1]
        
                    if second_column_value is None:
                        process_data = get_process_info(process_id) 
                        uwxid = process_data.get('UWXId', None) # Extract UWXId and store it in uwxid variable
                        user_active = is_user_active(uwxid)
                        
                        if not user_active:
                            fill_excel_and_skip_process_openpyxl(sheet, index)
                        fill_in_excel_openpyxl(sheet, index, process_data)
        
                        print(f"{index + 2}: {'Active' if user_active else 'Inactive'}, {process_id}", {process_data.get('ContractId', None)[:8]}, {process_data.get('PlacementId', None)[:8]}, {process_data.get('SectionId', None)[:8]}, {process_data.get('SubmissionId', None)[:8]}, {process_data.get('UWXId', None)[:8]})
                            
            finally:
                wb.save(excel_file)
                print("Excel saved")
                keyboard.remove_hotkey('ctrl+alt+s')
                driver.quit()