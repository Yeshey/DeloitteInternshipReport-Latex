        from urllib.parse import urljoin
        import json
        import pyautogui
        import pyperclip
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from bs4 import BeautifulSoup
        import re
        import time
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import pandas as pd
        from selenium.common.exceptions import WebDriverException
        from selenium.common.exceptions import WebDriverException, StaleElementReferenceException
        from selenium.webdriver.common.alert import Alert
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        import subprocess
        import keyboard
        import toml
        
        # pip install -r requirements.txt
        
        running = True
        
        def stop_script():
            global running
            print("Script stopping on next iteration")
            running = False
        
        def install_dependencies():
            with open('requirements.txt', 'r') as file:
                dependencies = file.read().splitlines()
        
            for dependency in dependencies:
                subprocess.run(['pip', 'install', dependency])
        
        def find_element(driver, xpath, timeout=10, max_attempts=5):
            button = None
            try:
                button = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            except (WebDriverException, StaleElementReferenceException) as e:
                print(f"Catching WebDriverException: {e}")
                button = None
        
            attempts = 0
            while attempts < max_attempts and button is None:
                try:
                    button = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                except (WebDriverException, StaleElementReferenceException):
                    attempts += 1
                    pass
            else:
                if attempts == max_attempts and button is None:
                    raise Exception(f"Failed after {max_attempts} attempts")
            return button
        
        def initialize_driver():
            global driver
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
        
            driver = webdriver.Chrome(options)  # options
        
            with open('credentials.toml', 'r') as file:
                credentials = toml.load(file)['credentials']
        
            process_id = "4850421"
            first_url = f"https://www.ril.com/ServiceCenter/Process_Instance_Detail.aspx?ProcessId={process_id}"
            driver.get(first_url)
        
            second_url = "https://www.ril.com/ServiceCenter/Error_Logs.aspx"
        
            # login into service center
            driver.switch_to.window(driver.window_handles[0])
            driver.implicitly_wait(3)
        
            # Find the username and password fields and fill them
            username_field = driver.find_element(By.ID, "wt89_wtContentRight_wtInput1")
            password_field = driver.find_element(By.ID, "wt89_wtContentRight_wtInputPass1")
        
            username_field.send_keys(credentials['username'])
            password_field.send_keys(credentials['password'])
        
            # Find and click the login button
            login_button = driver.find_element(By.ID, "wt89_wtContentRight_wt59_wtColumnsItems_wt33_wtContent_wtButton1")
            login_button.click()
        
            driver.execute_script(f"window.open('{second_url}', '_blank');")
            
        def get_error_and_url(dates):
            paint_this_row = False
            driver.switch_to.window(driver.window_handles[1])
            error = ""
            url = ""
            input_date_begin = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[2]/div[2]/div/div[1]/div/input") 
            input_date_begin.send_keys(Keys.CONTROL, 'a')
            input_date_begin.send_keys(Keys.DELETE)
            input_date_begin.send_keys(dates[1])
            input_date_end = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[2]/div[2]/div/div[2]/div/input") 
            input_date_end.send_keys(Keys.CONTROL, 'a')
            input_date_end.send_keys(Keys.DELETE)
            input_date_end.send_keys(dates[0])
            
            # dropdown
            dropdown = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div") 
            dropdown.click()
            
            dropdown_search = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/div/div[1]/div[2]/div[1]/input") 
            dropdown_search.send_keys("SDC_CS")
            dropdown_search.send_keys(Keys.ENTER)
            
            filter_btn = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[2]/div[2]/div/div[4]/div/div/input[1]") 
            filter_btn.click()
            
            time.sleep(2) # wait for it to load
            
            try:
                table = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/span[1]/div[2]/table/tbody") 
            except Exception as e:
                print(f"Exception: Couldn't find any errors? Painting row to view later")
                url = ""
                error = ""
                paint_this_row = True
            else:
                html_content = table.get_attribute('innerHTML')
                soup = BeautifulSoup(html_content, 'html.parser')
                tr_elements = soup.find_all('tr')
                for tr in tr_elements:
                    td_elements = tr.find_all('td')
                    message = td_elements[3].get_text(strip=True)
                    #print(f"td_elements:\n{td_elements}")
                    #print(f"message:\n{message}")
                    if "Could not execute a built-in/extended action on activity 'IsSuspended" in message:
                        continue
                    else:
                        error = td_elements[3].get_text(strip=True)
                        # get the url
                        for detail in td_elements:
                            if detail.find_all('a') and "Detail" in detail.get_text(strip=True):
                                anchor = detail.find('a')
                                if anchor and 'href' in anchor.attrs:
                                    relative_url = anchor['href']
                                    base_url = "https://www.ril.com/ServiceCenter/"
                                    url = urljoin(base_url, relative_url)
                                    # Now you can use the 'url' variable as needed in your code
                                else:
                                    raise Exception("Anchor element not found within detail element.")
                        break
                    
                if "Message: 'Leader " in error:
                    error = "Leader Error"
                elif "DUNS provided is not valid" in error:
                    error = "DUNS Number error"
                else:
                    paint_this_row = True
            return error, url, paint_this_row
            
        def get_process_last_error_window(process_id):
            driver.switch_to.window(driver.window_handles[0])
            url = f"https://www.ril.com/ServiceCenter/Process_Instance_Detail.aspx?ProcessId={process_id}"
            driver.get(url)
            driver.implicitly_wait(5)
            table = find_element(driver, "/html/body/form/div[3]/div[2]/div[1]/div[2]/div[3]/table/tbody") 
            html_content = table.get_attribute('innerHTML')
            soup = BeautifulSoup(html_content, 'html.parser')
            tr_elements = soup.find_all('tr')
            dates_list = []
            # Iterate over the first two <tr> elements
            for tr in tr_elements[:2]:
                td_elements = tr.find_all('td')
                # Check if there are at least 7 <td> elements
                if len(td_elements) >= 7:
                    # Extract the date from the seventh <td> element
                    date_text = td_elements[6].get_text(strip=True)
                    
                    # Append the date to the dates_list
                    dates_list.append(date_text)
            return dates_list
        
        def fill_excel_and_change_to_suspended(sheet, row_index, error, url, paint):
            # Update the specified columns in the row using openpyxl
            if not paint:
                sheet.cell(row=row_index + 2, column=7, value="Suspended")  # column for Comment
            sheet.cell(row=row_index + 2, column=11, value=error)  # Third column for PlacementId
            sheet.cell(row=row_index + 2, column=12, value=url)  # Third column for PlacementId
        
            # Check if the 'paint' variable is True, and if so, change the row color
            if paint:
                fill = PatternFill(start_color="d69cff", end_color="d69cff", fill_type="solid")  # 800080 is the hexadecimal code for purple
                for cell in sheet[row_index + 2]:
                    cell.fill = fill
        
        if __name__ == "__main__":
            install_dependencies()
            # Load the Excel file with openpyxl
            excel_file = 'SDC_processes.xlsx'
            wb = load_workbook(excel_file)
            sheet = wb['SDC_Generation']
            
            keyboard.add_hotkey('ctrl+alt+s', stop_script) # stop with CTRL + ALT + DELETE
            
            initialize_driver()
            try:
                # analyse ProcessIds in the first column from top to bottom if the second column is empty
                for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True)):
                    if (index + 2 > 1600):
                        # only analysing from the line 1600 onwards
                        process_id = row[0]
                        if process_id is None: # we reached the end
                            print("No more processIds to analyse")
                            break
                        if running == False:
                            print("script stopping...")
                            break
                        comment_column_value = row[6]
        
                        if comment_column_value == "Active - Loop":
                            dates = get_process_last_error_window(process_id)
                            error, url, paint_this_row = get_error_and_url(dates)
                            fill_excel_and_change_to_suspended(sheet, index, error, url, paint_this_row)
                            
                            print(f"{index+2}: {process_id}, newer date: {dates[0]}, older date: {dates[1]}, error: {error}, url: {url}")
                            
            finally:
                wb.save(excel_file)
                print("Excel saved")
                keyboard.remove_hotkey('ctrl+alt+s')
                driver.quit()